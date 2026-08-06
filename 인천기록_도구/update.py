# -*- coding: utf-8 -*-
"""
인천 2026 기록 갱신기
  python3 update.py payload.json
payload.json 스키마는 SKILL.md 참고. 모든 키는 선택(있는 것만 반영).
실행 순서: 워크북 갱신 → LibreOffice 재계산 → data.json 추출 → 대시보드 재생성
"""
import openpyxl, json, sys, os, shutil, subprocess, warnings, datetime
warnings.filterwarnings("ignore")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import charts as CH
from positions import compute_positions

BASE = os.path.dirname(os.path.abspath(__file__))          # 인천기록_도구
ROOT = os.path.dirname(BASE)                                # 기록 폴더
XLSX = os.path.join(ROOT, "인천 2026 기록.xlsx")
HTML = os.path.join(ROOT, "인천 2026 대시보드.html")
TPL  = os.path.join(BASE, "template.html")

pl = json.load(open(sys.argv[1], encoding="utf-8"))
wb = openpyxl.load_workbook(XLSX)
ws, ws2 = wb["2026 토탈"], wb["출전기록 (26)"]
log = []

# ── 1) 경기 기록 (2026 토탈 A~V) ──────────────────────────────
for m in pl.get("matches", []):
    comp, rnd = m.get("comp", "K리그"), m["round"]
    row = next((r for r in range(3, 42)
                if ws.cell(r, 1).value == comp and str(ws.cell(r, 4).value) == str(rnd)), None)
    if row is None:                       # FA컵 등 신규 행은 빈 행에 추가
        row = next(r for r in range(3, 60) if ws.cell(r, 1).value is None)
        ws.cell(row, 1, comp); ws.cell(row, 4, rnd)
    put = lambda c, v: ws.cell(row, c, v) if v is not None else None
    put(2, m.get("date")); put(3, m.get("dow")); put(5, m.get("opp")); put(6, m.get("ha"))
    put(7, m.get("gf"));   put(8, m.get("ga"))
    if m.get("gf") is not None and m.get("ga") is not None:
        ws.cell(row, 9, "승" if m["gf"] > m["ga"] else ("무" if m["gf"] == m["ga"] else "패"))
    put(10, m.get("tv")); put(11, m.get("commentator")); put(12, m.get("caster"))
    for i, s in enumerate(m.get("scorers", [])[:4]): ws.cell(row, 13 + i, s)
    for i, a in enumerate(m.get("assists", [])[:4]): ws.cell(row, 17 + i, a)
    put(21, m.get("att")); put(22, m.get("ref"))
    log.append(f"경기 {comp} {rnd}R → {row}행")

# ── 2) 출전기록 (선발 C:M / 교체 N:S) ─────────────────────────
for lu in pl.get("lineups", []):
    rnd = lu["round"]
    row = next((r for r in range(2, 40) if str(ws2.cell(r, 1).value) == str(rnd)), None)
    if row is None:
        log.append(f"!! 출전기록 {rnd}R 행 없음"); continue
    if lu.get("opp"): ws2.cell(row, 2, lu["opp"])
    for i, n in enumerate(lu.get("start", [])[:11]): ws2.cell(row, 3 + i, n)
    for i, n in enumerate(lu.get("sub", [])[:6]):    ws2.cell(row, 14 + i, n)
    log.append(f"출전기록 {rnd}R → {row}행 (선발 {len(lu.get('start',[]))} / 교체 {len(lu.get('sub',[]))})")

# ── 3) 서재민 베스트러너 (A42 표: 라운드 43행부터) ───────────
for rn in pl.get("runner", []):
    row = next((r for r in range(43, 76) if str(ws2.cell(r, 1).value) == str(rn["round"])), None)
    if row is None:
        log.append(f"!! 베스트러너 {rn['round']}R 행 없음"); continue
    if rn.get("opp"):  ws2.cell(row, 2, rn["opp"])
    if rn.get("rank") is not None: ws2.cell(row, 3, rn["rank"])
    if rn.get("dist") is not None: ws2.cell(row, 4, rn["dist"])
    res = rn.get("res") or next((ws.cell(r, 9).value for r in range(3, 41)
                                 if ws.cell(r, 1).value == "K리그"
                                 and str(ws.cell(r, 4).value) == str(rn["round"])), None)
    if res: ws2.cell(row, 5, res)
    log.append(f"베스트러너 {rn['round']}R → {row}행")

# ── 3.5) 리그 전체 결과 병합 → 순위 이력 재계산 ───────────────
import positions as POS
lr = POS.load()
for rd, games in (pl.get("league_results") or {}).items():
    lr[str(rd)] = [POS.parse_line(g) if isinstance(g, str) else g for g in games]
if pl.get("league_results"):
    POS.save(lr)
history, computed = POS.compute_positions(lr)
CH.rebuild(ws)
refs_seen = CH.rebuild_referees(ws)
log.append(f"순위 이력 {len(history['인천'])}R · 주심표 {len(refs_seen)}명")

# ── 4) 리그순위표 (BI3:BR14) ─────────────────────────────────
st = pl.get("standings") or (computed if lr else None)
if st:
    from openpyxl.styles import PatternFill, Font
    GOLD, ZONE = "E2B33C", {1: "D6E9FF", 2: "E9F3FF", 3: "E9F3FF", 12: "FFE3E0"}
    for i, row_data in enumerate(st[:12]):
        r = 3 + i
        pos, team, pld, w, d, l, gf, ga, pts = row_data
        for j, v in enumerate([pos, team, pld, w, d, l, gf, ga, f"=BO{r}-BP{r}", pts]):
            c = ws.cell(r, 61 + j, v)
            if team == "인천":
                c.fill = PatternFill("solid", fgColor=GOLD)
                c.font = Font(size=10, bold=True, color="1A1A1A")
            else:
                z = ZONE.get(pos)
                c.fill = PatternFill("solid", fgColor=z) if z else PatternFill(fill_type=None)
                c.font = Font(size=10)
    if pl.get("standings_asof"):
        ws["BI1"] = f"K리그1 2026 리그순위   ({pl['standings_asof']} 기준)"
    log.append(f"리그순위 {len(st)}팀 갱신")

wb.save(XLSX)

# ── 5) 재계산 → data.json → 대시보드 ──────────────────────────
TMP = "/tmp/_incheon"; shutil.rmtree(TMP, ignore_errors=True); os.makedirs(TMP)
shutil.copy(XLSX, f"{TMP}/t.xlsx")
subprocess.run(["libreoffice", "--headless", "--calc", "--convert-to",
                "xlsx:Calc MS Excel 2007 XML", "--outdir", f"{TMP}/o", f"{TMP}/t.xlsx"],
               capture_output=True, timeout=420)
w2 = openpyxl.load_workbook(f"{TMP}/o/t.xlsx", data_only=True)
s1, s2 = w2["2026 토탈"], w2["출전기록 (26)"]
cellv = lambda s, r, c: s.cell(r, c).value

matches = [{"comp": cellv(s1, r, 1), "date": cellv(s1, r, 2), "dow": cellv(s1, r, 3),
            "round": cellv(s1, r, 4), "opp": cellv(s1, r, 5), "ha": cellv(s1, r, 6),
            "gf": cellv(s1, r, 7), "ga": cellv(s1, r, 8), "res": cellv(s1, r, 9),
            "scorers": [cellv(s1, r, c) for c in range(13, 17) if cellv(s1, r, c)],
            "assists": [cellv(s1, r, c) for c in range(17, 21) if cellv(s1, r, c)],
            "att": cellv(s1, r, 21), "ref": cellv(s1, r, 22)}
           for r in range(3, 42) if cellv(s1, r, 1)]
players = [{"name": cellv(s1, r, 24), "app": cellv(s1, r, 25) or 0, "start": cellv(s1, r, 26) or 0,
            "sub": cellv(s1, r, 27) or 0, "g": cellv(s1, r, 28) or 0, "a": cellv(s1, r, 29) or 0,
            "ga": cellv(s1, r, 30) or 0} for r in range(3, 41) if cellv(s1, r, 24)]
split = [{"k": cellv(s1, r, 39), "pld": cellv(s1, r, 40), "w": cellv(s1, r, 41),
          "d": cellv(s1, r, 42), "l": cellv(s1, r, 43), "pts": cellv(s1, r, 44),
          "gf": cellv(s1, r, 46), "ga": cellv(s1, r, 47), "gd": cellv(s1, r, 48),
          "ppg": cellv(s1, r, 49)} for r in (3, 4, 5)]
table = [{"pos": cellv(s1, r, 61), "team": cellv(s1, r, 62), "pld": cellv(s1, r, 63),
          "w": cellv(s1, r, 64), "d": cellv(s1, r, 65), "l": cellv(s1, r, 66),
          "gf": cellv(s1, r, 67), "ga": cellv(s1, r, 68), "gd": cellv(s1, r, 69),
          "pts": cellv(s1, r, 70)} for r in range(3, 15) if cellv(s1, r, 62)]
refs = [{"name": cellv(s1, r, 53), "pld": cellv(s1, r, 54) or 0, "w": cellv(s1, r, 55) or 0,
         "d": cellv(s1, r, 56) or 0, "l": cellv(s1, r, 57) or 0, "pts": cellv(s1, r, 58) or 0}
        for r in range(3, 40)
        if cellv(s1, r, 53) and cellv(s1, r, 53) != "총합"]
att = {k: {"n": cellv(s1, rr, 14), "tot": cellv(s1, rr, 15), "avg": cellv(s1, rr, 16),
           "py_avg": cellv(s1, rr, 19), "yoy": cellv(s1, rr, 20)}
       for k, rr in (("kl", 44), ("fa", 45))}
runner = [{"round": cellv(s2, r, 1), "opp": cellv(s2, r, 2), "rank": cellv(s2, r, 3),
           "dist": cellv(s2, r, 4), "res": cellv(s2, r, 5)}
          for r in range(43, 76) if cellv(s2, r, 3) is not None]
lineups = [{"round": cellv(s2, r, 1), "opp": cellv(s2, r, 2),
            "start": [cellv(s2, r, c) for c in range(3, 14) if cellv(s2, r, c)],
            "sub": [cellv(s2, r, c) for c in range(14, 20) if cellv(s2, r, c)]}
           for r in range(2, 40) if cellv(s2, r, 1) is not None]

played = [m for m in matches if m["res"] and m["comp"] == "K리그"]
last = played[-1] if played else None
asof = pl.get("standings_asof") or (f"{last['date']} · {last['round']}R 종료" if last else "")
data = {"matches": matches, "players": players, "split": split, "table": table, "refs": refs,
        "att": att, "runner": runner, "lineups": lineups, "positions": history,
        "updated": datetime.date.today().isoformat(), "asof": asof}
json.dump(data, open(os.path.join(BASE, "data.json"), "w", encoding="utf-8"),
          ensure_ascii=False, indent=1)
tpl = open(TPL, encoding="utf-8").read()
page = tpl.replace("__DATA__", json.dumps(data, ensure_ascii=False))
open(HTML, "w", encoding="utf-8").write(page)
# GitHub Pages 용 index.html (같은 내용)
open(os.path.join(ROOT, "index.html"), "w", encoding="utf-8").write(page)

print("\n".join(log))
me = next((t for t in table if t["team"] == "인천"), None)
print(f"\n✔ 저장 완료\n  {XLSX}\n  {HTML}")
if me:
    print(f"  인천 {me['pos']}위 · 승점 {me['pts']} · {me['w']}승 {me['d']}무 {me['l']}패 "
          f"· {me['gf']}득 {me['ga']}실")
