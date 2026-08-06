# -*- coding: utf-8 -*-
"""워크북 정리 모듈 — update.py 에서 import.

엑셀 안의 차트는 렌더링이 깨지고 가독성이 나빠 전부 제거한다.
시각화는 HTML 대시보드가 전담한다. 워크북은 '리그순위표까지'만 유지.
"""
from openpyxl.utils import get_column_letter as gc

HELPER_COLS = range(72, 78)    # BT:BY  (구 차트 헬퍼)
POS_COLS = range(80, 93)       # CB:CN  (구 순위 이력 매트릭스)


def write_positions(ws, history):
    """(하위호환) 순위 이력은 이제 워크북에 쓰지 않는다. data.json 으로만 흐른다."""
    return len(history.get("인천", []))


def rebuild_referees(ws, max_rows=30):
    """V열(주심)에 새 이름이 나오면 BA 주심표에 자동으로 추가하고 총합 행을 옮긴다.
       BA=53 주심 / BB경기 BC승 BD무 BE패 BF승점 BG승률"""
    from copy import copy
    seen = []
    for r in range(3, 41):                      # K리그 구간만 집계 (원본 설계 유지)
        v = ws.cell(r, 22).value
        if v and v not in seen:
            seen.append(v)
    if not seen:
        return []

    proto = [ws.cell(3, 53 + i) for i in range(7)]   # 서식 견본 (기존 첫 행)
    styles = [(copy(c.font), copy(c.border), copy(c.fill),
               copy(c.alignment), c.number_format) for c in proto]

    for i, name in enumerate(seen):
        r = 3 + i
        ws.cell(r, 53, name)
        ws.cell(r, 54, f"=SUM(BC{r}:BE{r})")
        ws.cell(r, 55, f'=COUNTIFS($V$3:$V$40,$BA{r},$I$3:$I$40,BC$2)')
        ws.cell(r, 56, f'=COUNTIFS($V$3:$V$40,$BA{r},$I$3:$I$40,BD$2)')
        ws.cell(r, 57, f'=COUNTIFS($V$3:$V$40,$BA{r},$I$3:$I$40,BE$2)')
        ws.cell(r, 58, f"=BC{r}*3+BD{r}")
        ws.cell(r, 59, f"=IFERROR(BC{r}/BB{r},0)")
        for j, (fo, bo, fi, al, nf) in enumerate(styles):
            c = ws.cell(r, 53 + j)
            c.font, c.border, c.fill, c.alignment, c.number_format = fo, bo, fi, al, nf

    tot = 3 + len(seen)
    ws.cell(tot, 53, "총합")
    for col, letter in ((54, "BB"), (55, "BC"), (56, "BD"), (57, "BE"), (58, "BF")):
        ws.cell(tot, col, f"=SUM({letter}3:{letter}{tot - 1})")
    ws.cell(tot, 59, f"=IFERROR(BC{tot}/BB{tot},0)")
    for j, (fo, bo, fi, al, nf) in enumerate(styles):
        c = ws.cell(tot, 53 + j)
        bold = copy(fo); bold.b = True
        c.font, c.border, c.fill, c.alignment, c.number_format = bold, bo, fi, al, nf

    for r in range(tot + 1, tot + max_rows):     # 잔여 행 정리
        for col in range(53, 60):
            ws.cell(r, col).value = None
    return seen


def rebuild(ws):
    """차트를 모두 제거하고 차트 전용 헬퍼 영역을 비운다."""
    n = len(getattr(ws, "_charts", []))
    ws._charts = []
    for cols in (HELPER_COLS, POS_COLS):
        for c in cols:
            letter = gc(c)
            if letter in ws.column_dimensions:
                del ws.column_dimensions[letter]
            for r in range(1, 45):
                ws.cell(r, c).value = None
    return n
